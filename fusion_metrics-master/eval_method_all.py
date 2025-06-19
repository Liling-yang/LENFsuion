import os
import torch
import numpy as np
from PIL import Image
from natsort import natsorted
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from Metric import AG_function, CC_function, EN_function, MS_SSIM_function, SD_function, SF_function, VIF_function, MI_function, MSE_function, PSNR_function, SCD_function, Qabf_function, Nabf_function, SSIM_function, MS_SSIM_function

def my_round(x, decimals=3):
    if isinstance(x, torch.Tensor):
        return (x * 10**decimals).round() / 10**decimals
    else:
        return round(x, decimals)

def write_excel(excel_name, worksheet_name, column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
        workbook = Workbook()

    worksheet = workbook[worksheet_name] if worksheet_name in workbook.sheetnames else workbook.create_sheet(title=worksheet_name)

    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        if isinstance(value, torch.Tensor):
            value = value.item()
        worksheet[column + str(i + 1)].value = value

    workbook.save(excel_name)

def evaluation_one(ir_img, vi_img, f_img):
    EN = EN_function(f_img)
    MI = MI_function(np.array(ir_img).astype(np.int32), np.array(vi_img).astype(np.int32), np.array(f_img).astype(np.int32), gray_level=256)
    SF = SF_function(f_img)
    SD = SD_function(f_img)
    AG = AG_function(f_img)
    VIF = VIF_function(ir_img, vi_img, f_img)
    CC = CC_function(ir_img, vi_img, f_img)
    MSE = MSE_function(ir_img, vi_img, f_img)
    PSNR = PSNR_function(ir_img, vi_img, f_img)
    SCD = SCD_function(ir_img, vi_img, f_img)
    Qabf = Qabf_function(ir_img, vi_img, f_img)
    Nabf = Nabf_function(ir_img, vi_img, f_img)
    SSIM = SSIM_function(ir_img, vi_img, f_img)
    MS_SSIM = MS_SSIM_function(ir_img, vi_img, f_img)
    return EN, SF, SD, AG, MI, VIF, CC, SCD, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM

if __name__ == '__main__':
    dataroot = r'/mnt/data1/tool/fusion_metrics-master/datasets/'
    results_root = '/mnt/data1/tool/fusion_metrics-master/datasets/LLVIP/' # 数据集目录，子集有If, ir, vi
    data = 'LLVIP'
    ir_dir = os.path.join(dataroot, data, 'ir')
    vi_dir = os.path.join(dataroot, data, 'vi')

    Method_list = ['CoCoNet','CrossFuse','DDBF','LENFusion'] # 所对比的融合方法
    dataset = 'If'
    f_dir = os.path.join(results_root, dataset) # fused results目录

    dataname = 'LLVIP-test' # 存储数据的excel文件名
    save_dir = os.path.join(os.getcwd(), 'Metric') # 存储数据的excel文件的目录
    os.makedirs(save_dir, exist_ok=True)
    metric_save_name = os.path.join(save_dir, 'metric_{}.xlsx'.format(dataname)) # 存储数据的excel文件

    filelist = natsorted(os.listdir(ir_dir))
    
    for i, Method in enumerate(Method_list):
        EN_list,MI_list,SF_list,AG_list,SD_list,VIF_list,CC_list, SCD_list, MSE_list, PSNR_list, Qabf_list, Nabf_list, SSIM_list, MS_SSIM_list = ([] for _ in range(14))
        filename_list = ['']

        sub_f_dir = os.path.join(f_dir, Method)
        eval_bar = tqdm(filelist)
        for item in eval_bar:
            ir_img = np.array(Image.open(os.path.join(ir_dir, item)).convert('L')).astype(np.float32)
            vi_img = np.array(Image.open(os.path.join(vi_dir, item)).convert('L')).astype(np.float32)
            f_img = np.array(Image.open(os.path.join(sub_f_dir, item)).convert('L')).astype(np.float32)

            EN, SF, SD, AG, MI, VIF, CC, SCD, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM = evaluation_one(ir_img, vi_img, f_img)
            EN_list.append(EN)
            MI_list.append(MI)
            SF_list.append(SF)
            AG_list.append(AG)
            SD_list.append(SD)
            VIF_list.append(VIF)
            CC_list.append(CC)
            SCD_list.append(SCD)
            MSE_list.append(MSE)
            PSNR_list.append(PSNR)
            Qabf_list.append(Qabf)
            Nabf_list.append(Nabf)
            SSIM_list.append(SSIM)
            MS_SSIM_list.append(MS_SSIM)
            filename_list.append(item)
            eval_bar.set_description("{} | {}".format(Method, item))

        # 计算均值和标准差
        metrics_lists = [EN_list,MI_list,SF_list,AG_list,SD_list,VIF_list,CC_list, SCD_list, MSE_list, PSNR_list, Qabf_list, Nabf_list, SSIM_list, MS_SSIM_list]
        for metric_list in metrics_lists:
            metric_list.append(np.mean(metric_list))
            metric_list.append(np.std(metric_list))

        filename_list.append('mean')
        filename_list.append('std')

        metrics_lists = [[my_round(x) for x in metric_list] for metric_list in metrics_lists]

        for j, metric_list in enumerate(metrics_lists):
            metric_list.insert(0, Method)
            if i == 0:
                write_excel(metric_save_name, ['EN','MI','SF','AG','SD','VIF','CC', 'SCD', 'MSE', 'PSNR', 'Qabf', 'Nabf', 'SSIM', 'MS_SSIM'][j], 0, filename_list)
            write_excel(metric_save_name, ['EN','MI','SF','AG','SD','VIF','CC', 'SCD', 'MSE', 'PSNR', 'Qabf', 'Nabf', 'SSIM', 'MS_SSIM'][j], i + 1, metric_list)
